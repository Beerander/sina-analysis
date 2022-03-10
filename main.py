#coding:utf-8
import re
import time

import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook


def get_html(url):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Cookie': 'login_sid_t=bd7bbb25c0e36a32c8624aa10619baa9; cross_origin_proto=SSL; _s_tentry=passport.weibo.com; Apache=3392348386572.077.1646830776746; SINAGLOBAL=3392348386572.077.1646830776746; ULV=1646830776750:1:1:1:3392348386572.077.1646830776746:; SUB=_2A25PLNVFDeRhGeNK7FMS-SvNyz2IHXVsWEGNrDV8PUNbmtB-LW7wkW9NSVuz9kMSOgg80D9rVhszw3L3YgocjdTc; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5CCU5FaD5XSIYUvR5NlMqk5JpX5KzhUgL.Fo-XS0201K-peh22dJLoI7yKdNHXI2iL9Btt; ALF=1678366869; SSOLoginState=1646830869',
        'Host': 's.weibo.com',
        'Pragma': 'no-cache',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Microsoft Edge";v="99"',
        'sec-ch-ua-mobile': '?0',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36 Edg/99.0.1150.36'
    }
    res = requests.get(url,headers=headers)
    if res.status_code == 200:
        print('获取成功')
        return res.text
    else:
        print('失败')

def jianxi(res):
    data = []
    res = re.findall('<!--card-wrap-->(.*?)<!--/card-wrap-->',res,re.S)
    for r in res:
        xp = etree.HTML(r)
        n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content_full"]//text()')
        if len(n) == 0:
            n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content"]//text()')
        t = xp.xpath('//div[@class="content"]/p[@class="from"]/a[1]/text()')
        p = xp.xpath('//div[@class="card-act"]//li[2]/a/text()')[0]
        d = xp.xpath('//div[@class="card-act"]//li[3]/a/button/span[2]/text()')
        if d[0] == '赞':
            d = '0'
        elif len(d) != 0 :
            d = d[0]
        else :
            d = '0'
        p =re.findall('\d*',p)
        p = ''.join('%s' % r.split() for r in p).replace('[', '').replace(']', '').replace('\'', '')
        if p == '':
            p = '0'
        t = ''.join(t[0].split())
        n = ''.join('%s' %r.split() for r in n).replace('[','').replace(']','').replace('\'','')
        n = re.sub(r'\\u...','',n)
        n = re.sub(r'收起全文d','',n)
        data.append({'时间': t, '评论数': p, '点赞数': d, '内容': n})
    return data

def write_data(datas):
    wb = load_workbook('南京农业大学相关微博.xlsx')
    ws = wb.create_sheet('南京农业大学相关微博', 0)
    ys = {
        'A':'时间',
        'B':'评论数',
        'C':'点赞数',
        'D':'内容'
    }
    for key, value in ys.items():
        ws[key + '1'] = value
    b = 0
    for data in datas:
        for n in range(len(list(data.values())[0])):
            for key, value in ys.items():
                ws[key + str(n + 2 + b)] = list(data.values())[0][n][value]
        b += len(list(data.values())[0])
    wb.save('南京农业大学相关微博.xlsx')

if __name__ == '__main__':
    wb = Workbook()
    wb.save('南京农业大学相关微博.xlsx')
    datas = []
    for i in range(1,11):
        url = 'https://s.weibo.com/weibo?q=%E5%8D%97%E4%BA%AC%E5%86%9C%E4%B8%9A%E5%A4%A7%E5%AD%A6&page='+str(i)
        res = get_html(url)
        data = jianxi(res)
        print(i,data)
        datas.append({str(i): data})
        time.sleep(0.5)
    write_data(datas)
